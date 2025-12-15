# toolkit/datatable.py
from __future__ import annotations

from typing import Any, Dict, List, Iterator, Tuple
from openpyxl import load_workbook

from toolkit.funlib import normalize


class SheetData:
    """
    封裝單一 Sheet 的資料列集合。
    每一列是一個 dict：{欄位名稱: 值}
    current_index 表示目前「游標」所在的列。
    """

    def __init__(self, rows: List[Dict[str, Any]]):
        self.rows = rows
        self.current_index = 0

    @property
    def row_count(self) -> int:
        return len(self.rows)

    @property
    def current_row(self) -> Dict[str, Any]:
        return self.rows[self.current_index]

    def get(self, col_name: str, default: Any | None = None) -> Any:
        return self.current_row.get(col_name, default)

    def set_current_row(self, index: int) -> None:
        if index < 0 or index >= len(self.rows):
            raise IndexError(f"索引 {index} 超出範圍，總列數為 {len(self.rows)}")
        self.current_index = index

    def get_current_row(self) -> int:
        return self.current_index

    def add_parameter(self, col_name: str, value: Any | None = None) -> None:
        for row in self.rows:
            if col_name not in row:
                row[col_name] = ""
        self.current_row[col_name] = value

    def set_row_by_value(self, col_name: str, value: Any) -> bool:
        original_index = self.current_index
        for index, row in enumerate(self.rows):
            if row.get(col_name) == value:
                self.current_index = index
                return True
        self.current_index = original_index
        return False


class DataTable:
    """
    管理多個 SheetData，提供類似 UFT DataTable 的操作體驗。
    以 alias 作為每個 Sheet 的識別名稱。
    """

    def __init__(self):
        self._sheets: Dict[str, SheetData] = {}

    def add_sheet_from_excel(self, alias: str, file_path: str, sheet_name: str) -> None:
        """
        從 Excel 載入指定 sheet，並以 alias 存入 DataTable。
        - 第一列視為欄位名稱
        - 第二列開始為資料列
        alias 重複載入将直接覆盖原本资料
        """
        wb = load_workbook(file_path, data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Excel 不存在分頁：'{sheet_name}' (file='{file_path}')")

        ws = wb[sheet_name]

        # 第一列當欄位名稱（normalize 防 None/空白）
        header_cells = next(ws.iter_rows(min_row=1, max_row=1))
        headers: List[str] = [normalize(c.value) for c in header_cells]

        # 欄位名稱防呆：不可空、不可重複
        seen: set[str] = set()
        for h in headers:
            if not h:
                raise ValueError(f"{sheet_name} sheet 錯誤：欄位名稱不可為空白")
            if h in seen:
                raise ValueError(f"{sheet_name} sheet 錯誤：欄位名稱重複：'{h}'")
            seen.add(h)

        rows: List[Dict[str, Any]] = []
        for row_cells in ws.iter_rows(min_row=2):
            data: Dict[str, Any] = {}
            for header, cell in zip(headers, row_cells):
                data[header] = cell.value
            rows.append(data)

        self._sheets[alias] = SheetData(rows)

    def get_sheet(self, sheet: str) -> SheetData:
        return self._sheets[sheet]

    def has_sheet(self, sheet_name: str) -> bool:
        return sheet_name in self._sheets

    def get_sheet_count(self) -> int:
        return len(self._sheets)

    def get_data(self, col_name: str, sheet: str) -> Any:
        return self._sheets[sheet].get(col_name)

    def set_current_row(self, sheet: str, index: int = 0) -> None:
        self._sheets[sheet].set_current_row(index)

    def iter_rows(self, sheet: str) -> Iterator[Tuple[int, Dict[str, Any]]]:
        sheet_data = self._sheets[sheet]
        for i, row in enumerate(sheet_data.rows):
            yield i, row
